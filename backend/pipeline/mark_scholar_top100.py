"""
Mark Google Scholar Top 100 most-cited papers as canonical.

Reads GoogleScholartop100.xlsx, for each paper:
1. Looks up in our DB by DOI (most are already there — top-cited!)
2. If not in DB, fetches from OpenAlex
3. Marks as canonical with field=canonical-gs-top100

Usage:
  python -m backend.pipeline.mark_scholar_top100 --budget 5
"""

import argparse
import asyncio
import json
import os
import re
import sys
from datetime import datetime, timezone

import httpx
import openpyxl
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
OPENALEX_BASE = "https://api.openalex.org"
ANTHROPIC_API = "https://api.anthropic.com/v1/messages"

HAIKU_INPUT = 0.80 / 1_000_000
HAIKU_OUTPUT = 4.0 / 1_000_000

budget_spent = 0.0
budget_limit = 5.0
_lock = None

HEADERS_SUPABASE = {
    "apikey": SUPABASE_SERVICE_KEY,
    "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

from backend.prompts.paper_analysis import ANALYSIS_PROMPT


async def supabase_get(client, table, params):
    r = await client.get(
        f"{SUPABASE_URL}/rest/v1/{table}", params=params,
        headers={**HEADERS_SUPABASE, "Prefer": ""}, timeout=15,
    )
    return r.json() if r.status_code == 200 else []


async def supabase_post(client, table, data):
    r = await client.post(
        f"{SUPABASE_URL}/rest/v1/{table}",
        json=data, headers=HEADERS_SUPABASE, timeout=15,
    )
    if r.status_code in (200, 201):
        return r.json()
    return None


async def supabase_patch(client, table, params, data):
    r = await client.patch(
        f"{SUPABASE_URL}/rest/v1/{table}", params=params,
        json=data, headers=HEADERS_SUPABASE, timeout=15,
    )
    return r.status_code in (200, 204)


def reconstruct_abstract(inv):
    if not inv:
        return ""
    pos = []
    for word, poss in inv.items():
        for p in poss:
            pos.append((p, word))
    pos.sort()
    return " ".join(w for _, w in pos)


async def find_paper_in_db(client, doi, title):
    """Check if paper is already in our DB."""
    # Try by DOI first
    if doi:
        clean_doi = doi.replace("https://doi.org/", "")
        # Search variations
        for pattern in [f"eq.https://doi.org/{clean_doi}", f"eq.{clean_doi}"]:
            results = await supabase_get(client, "papers", {
                "doi": pattern, "select": "id,title,canonical",
            })
            if results:
                return results[0]
    # Try by title
    title_short = title[:50].replace("'", "").replace(":", "").replace("&", "")
    results = await supabase_get(client, "papers", {
        "title": f"ilike.*{title_short}*", "select": "id,title,canonical", "limit": "3",
    })
    for r in results:
        if r["title"].lower()[:30] == title.lower()[:30]:
            return r
    return results[0] if results else None


async def search_openalex_by_doi_or_title(client, doi, title):
    """Fetch paper metadata from OpenAlex."""
    try:
        if doi:
            clean_doi = doi.replace("https://doi.org/", "")
            r = await client.get(f"{OPENALEX_BASE}/works/doi:{clean_doi}", params={
                "select": "id,title,authorships,publication_year,abstract_inverted_index,cited_by_count,doi,primary_location",
            }, timeout=15)
            if r.status_code == 200:
                return r.json()
        r = await client.get(f"{OPENALEX_BASE}/works", params={
            "search": title,
            "per_page": 3,
            "select": "id,title,authorships,publication_year,abstract_inverted_index,cited_by_count,doi,primary_location",
        }, timeout=15)
        if r.status_code == 200:
            results = r.json().get("results", [])
            if results:
                return results[0]
    except Exception:
        pass
    return None


async def analyze_abstract(client, title, authors, year, abstract):
    global budget_spent
    if not abstract or len(abstract) < 50:
        return None
    async with _lock:
        if budget_spent >= budget_limit:
            return "STOP"

    prompt = ANALYSIS_PROMPT.format(title=title, authors=authors, year=year or 0, abstract=abstract)
    body = {
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 2500, "temperature": 0.2,
        "messages": [{"role": "user", "content": prompt}],
    }
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "content-type": "application/json",
        "anthropic-version": "2023-06-01",
    }
    try:
        r = await client.post(ANTHROPIC_API, json=body, headers=headers, timeout=60)
        if r.status_code != 200:
            return None
        data = r.json()
        text = data["content"][0]["text"]
        usage = data.get("usage", {})
        cost = usage.get("input_tokens", 0) * HAIKU_INPUT + usage.get("output_tokens", 0) * HAIKU_OUTPUT
        async with _lock:
            budget_spent += cost
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            text = text.split("```")[1].split("```")[0]
        return json.loads(text.strip())
    except Exception:
        return None


async def process_paper(client, row, stats):
    rank = row[0]
    authors = row[1]
    title = row[2]
    journal = row[3]
    year = row[6]
    citations = row[7]
    doi = row[8]

    if not title:
        return

    # Check DB
    existing = await find_paper_in_db(client, doi, title)
    if existing:
        # Mark as canonical
        await supabase_patch(client, "papers", {"id": f"eq.{existing['id']}"}, {
            "canonical": True,
            "canonical_field": "canonical-gs-top100",
            "canonical_rank": rank,
            "canonical_reason": f"Google Scholar Top 100 most-cited papers (rank #{rank}, {citations:,} citations)",
        })
        # Upsert canonical_works entry
        cw_existing = await supabase_get(client, "canonical_works", {
            "field": "eq.canonical-gs-top100", "title": f"eq.{title}", "select": "id",
        })
        if not cw_existing:
            await supabase_post(client, "canonical_works", {
                "field": "canonical-gs-top100", "title": title, "author": authors,
                "year": year, "rank": rank, "paper_id": existing["id"],
                "why": f"Top 100 most-cited paper per Google Scholar ({citations:,} citations)",
            })
        else:
            await supabase_patch(client, "canonical_works", {"id": f"eq.{cw_existing[0]['id']}"}, {
                "paper_id": existing["id"],
            })
        stats["db_marked"] += 1
        print(f"  ✓ #{rank:3} [DB, {citations:,} cites] {title[:60]}")
        return

    # Not in DB — fetch from OpenAlex
    raw = await search_openalex_by_doi_or_title(client, doi, title)
    if not raw:
        stats["not_found"] += 1
        print(f"  ✗ #{rank:3} [NOT IN OA] {title[:60]}")
        return

    oa_id = raw["id"].split("/")[-1]
    abstract = reconstruct_abstract(raw.get("abstract_inverted_index"))

    # Insert with canonical flag
    authorships = raw.get("authorships", [])
    authors_json = json.dumps([
        {"name": a.get("author", {}).get("display_name", "")} for a in authorships[:10]
    ])
    paper_row = {
        "openalex_id": oa_id,
        "doi": raw.get("doi") or (doi if doi else None),
        "title": raw.get("title", title),
        "authors": authors_json,
        "publication_year": raw.get("publication_year") or year,
        "abstract": abstract,
        "source_journal": journal or ((raw.get("primary_location") or {}).get("source") or {}).get("display_name"),
        "cited_by_count": raw.get("cited_by_count", citations or 0),
        "canonical": True,
        "canonical_field": "canonical-gs-top100",
        "canonical_rank": rank,
        "canonical_reason": f"Google Scholar Top 100 most-cited papers (rank #{rank}, {citations:,} citations)",
        "is_stub": (not abstract or len(abstract) < 50),
    }

    if abstract and len(abstract) > 50:
        analysis = await analyze_abstract(client, title, authors or "", year, abstract)
        if analysis and not isinstance(analysis, str):
            pt = analysis.get("paper_type") or {}
            paper_row["paper_type"] = pt.get("type")
            paper_row["subfield"] = pt.get("subfield")
            paper_row["analysis"] = json.dumps(analysis)
            paper_row["analysis_model"] = "claude-haiku-4-5-20251001"
            paper_row["analyzed_at"] = datetime.now(tz=timezone.utc).isoformat()
            paper_row["is_stub"] = False

    result = await supabase_post(client, "papers", paper_row)
    if result:
        paper_id = result[0]["id"]
        await supabase_post(client, "canonical_works", {
            "field": "canonical-gs-top100", "title": title, "author": authors,
            "year": year, "rank": rank, "paper_id": paper_id,
            "why": f"Top 100 most-cited paper per Google Scholar ({citations:,} citations)",
        })
        stats["inserted"] += 1
        print(f"  + #{rank:3} [NEW, {citations:,} cites] {title[:60]}")


async def async_main():
    global budget_limit, _lock
    sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default="GoogleScholartop100.xlsx")
    parser.add_argument("--budget", type=float, default=5.0)
    args = parser.parse_args()

    budget_limit = args.budget
    _lock = asyncio.Lock()

    wb = openpyxl.load_workbook(args.file)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    print(f"\n{'='*60}")
    print(f"GOOGLE SCHOLAR TOP 100 CANONICAL MARKING")
    print(f"Rows: {len(rows)} | Budget: ${budget_limit}")
    print(f"{'='*60}\n")

    stats = {"db_marked": 0, "inserted": 0, "not_found": 0}
    async with httpx.AsyncClient() as client:
        for row in rows:
            if budget_spent >= budget_limit:
                print("\n*** BUDGET LIMIT ***")
                break
            if not row[2]:
                continue
            try:
                await process_paper(client, row, stats)
            except Exception as e:
                print(f"  ERROR on #{row[0]}: {e}")

    print(f"\n{'='*60}")
    print(f"RESULTS:")
    print(f"  Already in DB (marked): {stats['db_marked']}")
    print(f"  New papers inserted:    {stats['inserted']}")
    print(f"  Not found:              {stats['not_found']}")
    print(f"  Cost:                   ${budget_spent:.3f}")
    print(f"{'='*60}")


def main():
    asyncio.run(async_main())


if __name__ == "__main__":
    main()
